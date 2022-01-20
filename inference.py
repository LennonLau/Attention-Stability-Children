import torch
import numpy as np
from network import C3D_model
import cv2
import os
from openpyxl import load_workbook, Workbook
from tqdm import tqdm
torch.backends.cudnn.benchmark = True

def CenterCrop(frame, size):
    h, w = np.shape(frame)[0:2]
    th, tw = size
    x1 = int(round((w - tw) / 2.))
    y1 = int(round((h - th) / 2.))

    frame = frame[y1:y1 + th, x1:x1 + tw, :]
    return np.array(frame).astype(np.uint8)


def center_crop(frame):
    frame = frame[8:120, 30:142, :]
    return np.array(frame).astype(np.uint8)


def main():
    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    print("Device being used:", device)

    # with open('./dataloaders/ucf_labels.txt', 'r') as f:
    #     class_names = f.readlines()
    #     f.close()
    # init model
    model = C3D_model.C3D(num_classes=2)
    checkpoint = torch.load(r'run/run_9/models/C3D-360_epoch-99.pth.tar', map_location=lambda storage, loc: storage)
    model.load_state_dict(checkpoint['state_dict'])
    model.to(device)
    model.eval()

    # read video
    pth_videos = r'data/360-label01test'
    if not os.path.exists('inference_test.xlsx'):
        wb = Workbook('inference_test.xlsx')
        wb.create_sheet('inference_test_sheet')
        wb.save('inference_test.xlsx')
    wb = load_workbook('inference_test.xlsx')
    ws = wb[wb.sheetnames[0]]
    ws.cell(1, 1).value = 'videoname'
    ws.cell(1, 2).value = 'real_label'
    ws.cell(1, 3).value = 'inference_label'
    video_nums = 0
    for label in os.listdir(pth_videos):
        for video in tqdm(os.listdir(os.path.join(pth_videos, label))):
            video_nums += 1
            label_flag = 1
            cap = cv2.VideoCapture(os.path.join(pth_videos, label, video))
            retaining = True

            clip = []
            while retaining:
                retaining, frame = cap.read()
                if not retaining and frame is None:
                    continue
                tmp_ = center_crop(cv2.resize(frame, (171, 128)))
                tmp = tmp_ - np.array([[[90.0, 98.0, 102.0]]])
                clip.append(tmp)
                if len(clip) == 16:
                    inputs = np.array(clip).astype(np.float32)
                    inputs = np.expand_dims(inputs, axis=0)
                    inputs = np.transpose(inputs, (0, 4, 1, 2, 3))
                    inputs = torch.from_numpy(inputs)
                    inputs = torch.autograd.Variable(inputs, requires_grad=False).to(device)
                    with torch.no_grad():
                        outputs = model.forward(inputs)

                    probs = torch.nn.Softmax(dim=1)(outputs)
                    infer_label = torch.max(probs, 1)[1].detach().cpu().numpy()[0]

                    if infer_label == 0:  # label == 0 means 'label0', label == 1 means 'label1'
                        label_flag = 0
                        break
                    clip.pop(0)

            ws.cell(video_nums + 1, 1).value = video
            ws.cell(video_nums + 1, 2).value = label
            ws.cell(video_nums + 1, 3).value = 'label0' if label_flag == 0 else 'label1'
            cap.release()

    wb.save('inference_test.xlsx')

if __name__ == '__main__':
    main()









