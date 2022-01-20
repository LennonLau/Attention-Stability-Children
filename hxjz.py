# -*-coding:utf-8-*-
from sklearn.metrics import confusion_matrix
import matplotlib.pyplot as plt
import numpy as np
import argparse
import random
import os
import shutil
from tqdm import tqdm
from openpyxl import load_workbook

labels = ['label0', 'label1']


wb = load_workbook('inference_test.xlsx')
ws = wb[wb.sheetnames[0]]
rows = ws.max_row

f_real = open('re_label.txt', 'w')
f_infer = open('pr_label.txt', 'w')

for line in tqdm(range(2, rows + 1)):
    re_label = (ws.cell(line, 2).value[-1])
    pr_label = (ws.cell(line, 3).value[-1])
    f_real.write(re_label + '\n')
    f_infer.write(pr_label + '\n')

f_real.close()
f_infer.close()

y_true = np.loadtxt('re_label.txt')
y_pred = np.loadtxt('pr_label.txt')

tick_marks = np.array(range(len(labels))) + 0.5


def plot_confusion_matrix(cm, title='Confusion Matrix', cmap=plt.cm.binary):
    plt.imshow(cm, interpolation='nearest', cmap=cmap)
    plt.title(title)
    plt.colorbar()
    xlocations = np.array(range(len(labels)))
    plt.xticks(xlocations, labels, rotation=90)
    plt.yticks(xlocations, labels)
    plt.ylabel('True label')
    plt.xlabel('Predicted label')


cm = confusion_matrix(y_true, y_pred)
np.set_printoptions(precision=2)
# cm_normalized = cm.astype('float') / cm.sum(axis=1)[:, np.newaxis]
cm_normalized = cm.astype('float')
print(cm_normalized)
plt.figure(figsize=(12, 8), dpi=120)

ind_array = np.arange(len(labels))
x, y = np.meshgrid(ind_array, ind_array)

for x_val, y_val in zip(x.flatten(), y.flatten()):
    c = cm_normalized[y_val][x_val]
    if c > 0.01:
        plt.text(x_val, y_val, "%0.2f" % (c,), color='red', fontsize=16, va='center', ha='center')
# offset the tick
plt.gca().set_xticks(tick_marks, minor=True)
plt.gca().set_yticks(tick_marks, minor=True)
plt.gca().xaxis.set_ticks_position('none')
plt.gca().yaxis.set_ticks_position('none')
plt.grid(True, which='minor', linestyle='-')
plt.gcf().subplots_adjust(bottom=0.15)

plot_confusion_matrix(cm_normalized, title='Normalized confusion matrix')
# show confusion matrix
plt.savefig('confusion_matrix.png', format='png')
plt.show()